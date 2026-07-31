from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="FF092B49"; GOLD="FFCEB684"; CREAM="FFF9F7F3"
BLUE=Font(name="Arial", size=10, color="FF0000FF")
BLK=Font(name="Arial", size=10)
BOLD=Font(name="Arial", size=10, bold=True)
HDR=Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
TITLE=Font(name="Arial", size=14, bold=True, color=NAVY)
SUB=Font(name="Arial", size=9, italic=True, color="FF5A6E84")
YEL=PatternFill("solid", fgColor="FFFFFF00")
NAVYF=PatternFill("solid", fgColor=NAVY)
GOLDF=PatternFill("solid", fgColor=GOLD)
CREAMF=PatternFill("solid", fgColor=CREAM)
thin=Side(style="thin", color="FFD8CDB8")
BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
CUR='$#,##0;($#,##0);"-"'
PCT='0.0%'

wb=Wb=Workbook()
ws=wb.active; ws.title="Offset Calculator"

def put(cell,val,font=BLK,fmt=None,fill=None,border=False,align=None):
    c=ws[cell]; c.value=val; c.font=font
    if fmt: c.number_format=fmt
    if fill: c.fill=fill
    if border: c.border=BOX
    if align: c.alignment=Alignment(horizontal=align)
    return c

ws.column_dimensions["A"].width=42
for col in "BCDE": ws.column_dimensions[col].width=16
ws.column_dimensions["F"].width=44

put("A1","TitanOS — per-family price and property-management offset",TITLE)
put("A2","Fill the yellow cells for one prospect. Everything else calculates. "
         "Blue text = an input you may change; black = a formula.",SUB)

# ── Inputs ────────────────────────────────────────────────────────────────
put("A4","THE HOUSEHOLD",HDR,fill=NAVYF); put("B4","",fill=NAVYF); put("C4","",fill=NAVYF)
rows=[("Household name","Prospect name here",None),
      ("Properties",6,'0'),
      ("Entities (LLCs, trusts)",3,'0'),
      ("Gross rental income — MONTHLY",25000,CUR)]
for i,(lab,val,fmt) in enumerate(rows, start=5):
    put(f"A{i}",lab,BLK,border=True)
    put(f"B{i}",val,BLUE,fmt,YEL,border=True)
put("C8","← monthly, not annual",SUB)
put("A9","Gross rental income — annual",BOLD,border=True)
put("B9","=B8*12",BOLD,CUR,border=True)

put("A11","WHAT THEY PAY A MANAGER TODAY",HDR,fill=NAVYF)
put("B11","",fill=NAVYF); put("C11","",fill=NAVYF)
put("A12","Management fee % of gross rent",BLK,border=True)
put("B12",0.10,BLUE,PCT,YEL,border=True)
put("C12","Market: 8–12%, 10% typical",SUB)
put("A13","Leasing + renewal + maintenance markup, as % of gross",BLK,border=True)
put("B13",0.05,BLUE,PCT,YEL,border=True)
put("C13","Industry all-in figure is 15–20% incl. this",SUB)
put("A14","Annual fee at the headline rate",BLK,border=True)
put("B14","=B9*B12",BLK,CUR,border=True)
put("A15","All-in annual cost of outside management",BOLD,border=True)
put("B15","=B9*(B12+B13)",BOLD,CUR,GOLDF,border=True)

# ── Tier pricing ──────────────────────────────────────────────────────────
put("A17","TITAN PRICING",HDR,fill=NAVYF)
for col,lab in zip("BCD",("Core","Private","Estate")):
    put(f"{col}17",lab,HDR,fill=NAVYF,align="center")
put("E17","",fill=NAVYF)

put("A18","Monthly fee",BLK,border=True)
for col,v in zip("BCD",(750,2500,5000)): put(f"{col}18",v,BLUE,CUR,YEL,border=True)
put("A19","Onboarding (one-off)",BLK,border=True)
for col,v in zip("BCD",(3500,7500,15000)): put(f"{col}19",v,BLUE,CUR,YEL,border=True)
put("A20","Properties included",BLK,border=True)
for col,v in zip("BCD",(3,6,12)): put(f"{col}20",v,BLUE,'0',YEL,border=True)
put("A21","Entities included",BLK,border=True)
for col,v in zip("BCD",(1,3,6)): put(f"{col}21",v,BLUE,'0',YEL,border=True)
put("A22","Per extra property, per month",BLK,border=True)
for col in "BCD": put(f"{col}22",200,BLUE,CUR,YEL,border=True)
put("A23","Per extra entity, per month",BLK,border=True)
for col in "BCD": put(f"{col}23",150,BLUE,CUR,YEL,border=True)

put("A25","QUOTE FOR THIS HOUSEHOLD",HDR,fill=NAVYF)
for col in "BCDE": put(f"{col}25","",fill=NAVYF)
put("A26","Base fee, annual",BLK,border=True)
put("A27","Property overage, annual",BLK,border=True)
put("A28","Entity overage, annual",BLK,border=True)
put("A29","Total annual fee",BOLD,border=True)
put("A30","Year one, including onboarding",BOLD,border=True)
for col in "BCD":
    put(f"{col}26",f"={col}18*12",BLK,CUR,border=True)
    # MAX guards the overage so an under-allowance household never shows a credit.
    put(f"{col}27",f"=MAX(0,$B$6-{col}20)*{col}22*12",BLK,CUR,border=True)
    put(f"{col}28",f"=MAX(0,$B$7-{col}21)*{col}23*12",BLK,CUR,border=True)
    put(f"{col}29",f"=SUM({col}26:{col}28)",BOLD,CUR,CREAMF,border=True)
    put(f"{col}30",f"={col}29+{col}19",BOLD,CUR,CREAMF,border=True)

put("A32","THE OFFSET",HDR,fill=NAVYF)
for col in "BCDE": put(f"{col}32","",fill=NAVYF)
put("A33","Outside management displaced (all-in)",BLK,border=True)
put("A34","Net annual cost after the offset",BOLD,border=True)
put("A35","Covered by the offset alone?",BOLD,border=True)
put("A36","Offset as a share of our fee",BLK,border=True)
for col in "BCD":
    put(f"{col}33","=$B$15",BLK,CUR,border=True)
    put(f"{col}34",f"={col}29-{col}33",BOLD,CUR,border=True)
    put(f"{col}35",f'=IF({col}33>={col}29,"Yes","No")',BOLD,None,None,border=True,align="center")
    # IFERROR guards a household with no rental income at all.
    put(f"{col}36",f"=IFERROR({col}33/{col}29,0)",BLK,PCT,border=True)

put("A38","BREAKEVEN — the gross rent at which each tier pays for itself",HDR,fill=NAVYF)
for col in "BCDE": put(f"{col}38","",fill=NAVYF)
put("A39","At the headline management rate only",BLK,border=True)
put("A40","At the all-in rate",BLK,border=True)
for col in "BCD":
    put(f"{col}39",f"=IFERROR({col}29/$B$12,0)",BLK,CUR,border=True)
    put(f"{col}40",f"=IFERROR({col}29/($B$12+$B$13),0)",BLK,CUR,border=True)

# ── Notes ────────────────────────────────────────────────────────────────
notes=[
 "HOW TO READ THIS",
 "",
 "Fill the yellow cells. Row 8 is MONTHLY rent — the single most likely",
 "input error, and it moves every figure below by 12x.",
 "",
 "BEFORE YOU USE THE OFFSET IN A PROPOSAL",
 "",
 "Confirm the family actually pays an outside manager. Across PCM's four",
 "households the only management fee on file is one Lamb property at 10%",
 "($3,000/yr); all six Kilcoyne properties record 0%, on $677,616 of",
 "gross rent. If a family self-manages there is no fee to displace, and",
 "the pitch is 'we take this off your desk', not 'we save you money'.",
 "",
 "Confirm what PCM actually replaces. Property management is a service:",
 "rent collection, tenant calls, maintenance dispatch, turnover. If PCM",
 "does oversight and coordination but not those, the family cannot fire",
 "their manager and the honest claim is a PARTIAL offset. Set row 12 to",
 "the portion genuinely displaced rather than the full rate.",
 "",
 "SOURCES FOR THE BENCHMARKS",
 "",
 "8–12% of gross rent, 10% typical for single-family, ~8.5% residential",
 "national average; all-in 15–20% once leasing (50–100% of first month's",
 "rent) and maintenance markups are counted.",
 "  home365.co/general-knowledge/property-management/",
 "    property-management-fees-2026/",
 "  allpropertymanagement.com/resources/ask-a-pro/posts/",
 "    how-much-property-managers-charge/",
 "  thepropertyceo.com/blog/property-management-fees-guide",
 "",
 "Default tier prices come from docs/PRICING-per-family.md; the market",
 "comparison and unit economics behind them are in that memo.",
]
for i,line in enumerate(notes, start=4):
    c=put(f"F{i}",line, BOLD if line.isupper() and line else SUB)
    c.alignment=Alignment(vertical="top")

put("A42","Example row above is a realistic household (6 properties, 3 entities, "
          "$25,000/month rent) — replace with the prospect's own figures.",SUB)

ws.sheet_view.showGridLines=False
ws.freeze_panes="A4"
wb.save("/sessions/epic-upbeat-volta/mnt/outputs/TitanOS_Pricing_Calculator.xlsx")
print("written")
